#!/usr/bin/env python3
"""Create Excel file with SA Large Caps calculation breakdown."""

import os
import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Load secrets
secrets_path = Path(__file__).parent.parent / '.streamlit' / 'secrets.toml'
if secrets_path.exists():
    try:
        import tomllib
    except ImportError:
        import tomli as tomllib
    with open(secrets_path, 'rb') as f:
        secrets = tomllib.load(f)
    for key, value in secrets.items():
        if key not in os.environ and isinstance(value, str):
            os.environ[key] = value

import pandas as pd
import numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from AlphaMachine_core.tracking import get_tracker
from AlphaMachine_core.data_manager import StockDataManager

tracker = get_tracker()
dm = StockDataManager()

# Get SA Large Caps portfolio - December holdings
portfolio = tracker.get_portfolio_by_name('SA Large Caps_EqualWeight')
holdings = tracker.get_holdings(portfolio.id, date(2025, 12, 4))
tickers = sorted([h.ticker for h in holdings])
weights = {h.ticker: float(h.weight) for h in holdings}

# Get December price data (include Nov 28 for first return calc)
price_data = dm.get_price_data(tickers, '2025-11-25', '2025-12-05')
df = pd.DataFrame(price_data)
df['trade_date'] = pd.to_datetime(df['trade_date']).dt.date

# Pivot to get tickers as columns
prices_df = df.pivot(index='trade_date', columns='ticker', values='close')
prices_df = prices_df.sort_index()
prices_df = prices_df[tickers]  # Sort columns

# Calculate daily returns
returns_df = prices_df.pct_change()

# Calculate weighted returns
weighted_returns_df = returns_df.copy()
for ticker in tickers:
    weighted_returns_df[ticker] = returns_df[ticker] * weights[ticker]

# Portfolio daily return
weighted_returns_df['Portfolio Return'] = weighted_returns_df[tickers].sum(axis=1)

# Create Excel workbook
wb = Workbook()

# Styling
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# === Sheet 1: Close Prices ===
ws1 = wb.active
ws1.title = 'Close Prices'

# Write weights row
ws1['A1'] = 'Weight'
ws1['A1'].font = Font(bold=True)
for col_idx, ticker in enumerate(tickers, start=2):
    cell = ws1.cell(row=1, column=col_idx, value=weights[ticker])
    cell.number_format = '0.0000'
    cell.font = Font(bold=True)

# Write headers
ws1['A2'] = 'Date'
ws1['A2'].font = header_font
ws1['A2'].fill = header_fill
for col_idx, ticker in enumerate(tickers, start=2):
    cell = ws1.cell(row=2, column=col_idx, value=ticker)
    cell.font = header_font
    cell.fill = header_fill

# Write price data
for row_idx, (trade_date, row) in enumerate(prices_df.iterrows(), start=3):
    ws1.cell(row=row_idx, column=1, value=trade_date)
    for col_idx, ticker in enumerate(tickers, start=2):
        cell = ws1.cell(row=row_idx, column=col_idx, value=row[ticker])
        cell.number_format = '0.00'

# === Sheet 2: Daily Returns ===
ws2 = wb.create_sheet('Daily Returns')

ws2['A1'] = 'Date'
ws2['A1'].font = header_font
ws2['A1'].fill = header_fill
for col_idx, ticker in enumerate(tickers, start=2):
    cell = ws2.cell(row=1, column=col_idx, value=ticker)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, (trade_date, row) in enumerate(returns_df.iterrows(), start=2):
    ws2.cell(row=row_idx, column=1, value=trade_date)
    for col_idx, ticker in enumerate(tickers, start=2):
        val = row[ticker]
        cell = ws2.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else None)
        cell.number_format = '0.00%'
        if pd.notna(val):
            cell.fill = green_fill if val >= 0 else red_fill

# === Sheet 3: Portfolio Calculation ===
ws3 = wb.create_sheet('Portfolio Calculation')

# Headers
headers = ['Date'] + tickers + ['Portfolio Return']
for col_idx, header in enumerate(headers, start=1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Row 2: Weights
ws3.cell(row=2, column=1, value='Weights')
ws3['A2'].font = Font(bold=True, italic=True)
for col_idx, ticker in enumerate(tickers, start=2):
    cell = ws3.cell(row=2, column=col_idx, value=weights[ticker])
    cell.number_format = '0.0000'
    cell.font = Font(italic=True)

# Data rows with weighted returns
for row_idx, (trade_date, row) in enumerate(weighted_returns_df.iterrows(), start=3):
    ws3.cell(row=row_idx, column=1, value=trade_date)
    for col_idx, ticker in enumerate(tickers, start=2):
        val = row[ticker]
        cell = ws3.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else None)
        cell.number_format = '0.0000%'
        if pd.notna(val):
            cell.fill = green_fill if val >= 0 else red_fill

    # Portfolio return
    port_ret = row['Portfolio Return']
    cell = ws3.cell(row=row_idx, column=len(tickers) + 2, value=port_ret if pd.notna(port_ret) else None)
    cell.number_format = '0.0000%'
    cell.font = Font(bold=True)
    if pd.notna(port_ret):
        cell.fill = green_fill if port_ret >= 0 else red_fill

# === Sheet 4: Summary ===
ws4 = wb.create_sheet('Summary')

ws4['A1'] = 'SA Large Caps - December 2025 Return Calculation'
ws4['A1'].font = Font(bold=True, size=14)

ws4['A3'] = 'Portfolio:'
ws4['B3'] = portfolio.name

ws4['A4'] = 'Holdings effective date:'
ws4['B4'] = '2025-12-01'

ws4['A5'] = 'Number of stocks:'
ws4['B5'] = len(tickers)

ws4['A6'] = 'Weight per stock:'
ws4['B6'] = 1/len(tickers)
ws4['B6'].number_format = '0.0000'

ws4['A8'] = 'Stocks:'
ws4['B8'] = ', '.join(tickers)

# December daily returns summary
ws4['A10'] = 'December Daily Returns'
ws4['A10'].font = Font(bold=True, size=12)

dec_returns = weighted_returns_df.loc[weighted_returns_df.index >= date(2025, 12, 1), 'Portfolio Return']
row_num = 11
for d, ret in dec_returns.items():
    ws4.cell(row=row_num, column=1, value=d)
    cell = ws4.cell(row=row_num, column=2, value=ret)
    cell.number_format = '0.0000%'
    row_num += 1

ws4.cell(row=row_num, column=1, value='December Total (compounded)')
ws4[f'A{row_num}'].font = Font(bold=True)
dec_total = (1 + dec_returns.dropna()).prod() - 1
cell = ws4.cell(row=row_num, column=2, value=dec_total)
cell.number_format = '0.0000%'
cell.font = Font(bold=True)

# Adjust column widths
for ws in [ws1, ws2, ws3, ws4]:
    ws.column_dimensions['A'].width = 12
    for i in range(2, 20):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        if i <= 26:
            ws.column_dimensions[col_letter].width = 10

ws4.column_dimensions['B'].width = 50

# Save
output_path = Path(__file__).parent.parent / 'SA_Large_Caps_Calculation.xlsx'
wb.save(output_path)
print(f'Excel file saved to: {output_path}')
print()
print('Sheets:')
print('  1. Close Prices - Raw closing prices with weights')
print('  2. Daily Returns - Daily % returns per stock')
print('  3. Portfolio Calculation - Weighted returns (return * weight) and portfolio total')
print('  4. Summary - Overview and December totals')
print()
print('December Returns Summary:')
for d, ret in dec_returns.items():
    print(f'  {d}: {ret*100:.4f}%')
print(f'  December Total (compounded): {dec_total*100:.4f}%')
