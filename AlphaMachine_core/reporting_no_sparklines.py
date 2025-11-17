import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from datetime import date
from scipy.stats import kurtosis, skew

def export_results_to_excel(engine, filepath):
    """Exportiert die Backtest-Ergebnisse als Excel-Datei (mit Trading-Kosten, Rebalance-Analyse und Next Month Allocation)."""
    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # … bestehende Blätter …
            if engine.selection_details:
                pd.DataFrame(engine.selection_details).to_excel(
                    writer, sheet_name="Selection Details", index=False
                )
            if engine.missing_months:
                pd.DataFrame(
                    {"Missing Rebalance Month": engine.missing_months}
                ).to_excel(writer, sheet_name="Missing Months", index=False)
            if engine.log_lines or engine.ticker_coverage_logs:
                pd.DataFrame(
                    {"Log Output": engine.ticker_coverage_logs + [""] + engine.log_lines}
                ).to_excel(writer, sheet_name="Run Log", index=False)
            if not engine.performance_metrics.empty:
                engine.performance_metrics.to_excel(
                    writer, sheet_name="Performance Summary", index=False
                )
            if not engine.daily_df.empty:
                engine.daily_df.to_excel(
                    writer, sheet_name="Daily Portfolio", index=False
                )
            if not engine.monthly_allocations.empty:
                engine.monthly_allocations.to_excel(
                    writer, sheet_name="Monthly Allocation", index=False
                )
            if not engine.monthly_performance.empty:
                engine.monthly_performance.to_excel(
                    writer, sheet_name="Monthly PnL", index=False
                )

            # Dashboard, Risk, Drawdowns, Trading Costs, Rebalance Analysis
            create_dashboard_sheet(engine, writer)
            create_risk_metrics_sheet(engine, writer)
            create_drawdown_sheet(engine, writer)
            create_trading_costs_sheet(engine, writer)
            create_rebalance_analysis_sheet(engine, writer)

            # —— neu: Next Month Allocation ——
            if hasattr(engine, 'next_month_weights'):
                df_next = engine.next_month_weights.mul(100).reset_index()
                df_next.columns = ['Ticker', 'Weight (%)']
                df_next.to_excel(writer,
                                 sheet_name='Next Month Allocation',
                                 index=False)

        # Portfolio-Chart ins Excel
        _add_portfolio_chart(filepath, engine.portfolio_value)

        print(f"✅ Excel-Report gespeichert unter: {filepath}")
        if engine.missing_months:
            print("⚠️ Fehlende Monate:", ", ".join(engine.missing_months))

    except Exception as e:
        print(f"❌ Fehler beim Export: {e}")


def _add_portfolio_chart(filepath, portfolio_series):
    """Fügt ein Liniendiagramm zur Excel-Datei hinzu."""
    try:
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(portfolio_series.index, portfolio_series.values, linewidth=2)
        ax.set_title("Portfolio Value Over Time")
        ax.set_xlabel("Date")
        ax.set_ylabel("Value")
        ax.grid(True)
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png")
        buf.seek(0)

        wb = load_workbook(filepath)
        ws = wb.create_sheet("Chart")
        img = XLImage(buf)
        img.anchor = "A1"
        ws.add_image(img)
        wb.save(filepath)
        buf.close()
    except Exception as e:
        print(f"❌ Fehler beim Hinzufügen des Charts: {e}")


def create_dashboard_sheet(engine, writer):
    """Erstellt ein einfaches Dashboard mit KPIs und Metadaten (inkl. Trading-Kosten und Rebalance-Frequenz)"""
    metrics = {
        "CAGR (%)": (
            engine.performance_metrics.loc[
                engine.performance_metrics["Metric"] == "CAGR (%)", "Value"
            ].values[0]
            if not engine.performance_metrics.empty
            else "n/a"
        ),
        "Sharpe Ratio": (
            engine.performance_metrics.loc[
                engine.performance_metrics["Metric"] == "Sharpe Ratio", "Value"
            ].values[0]
            if not engine.performance_metrics.empty
            else "n/a"
        ),
        "Max Drawdown (%)": (
            engine.performance_metrics.loc[
                engine.performance_metrics["Metric"] == "Max Drawdown (%)", "Value"
            ].values[0]
            if not engine.performance_metrics.empty
            else "n/a"
        ),
        "Start Date": (
            engine.portfolio_value.index.min().date()
            if not engine.portfolio_value.empty
            else "n/a"
        ),
        "End Date": (
            engine.portfolio_value.index.max().date()
            if not engine.portfolio_value.empty
            else "n/a"
        ),
        "Total Trading Costs": (
            engine.performance_metrics.loc[
                engine.performance_metrics["Metric"] == "Total Trading Costs", "Value"
            ].values[0]
            if not engine.performance_metrics.empty
            else "n/a"
        ),
        "Trading Costs (% of Initial)": (
            engine.performance_metrics.loc[
                engine.performance_metrics["Metric"] == "Trading Costs (% of Initial)",
                "Value",
            ].values[0]
            if not engine.performance_metrics.empty
            else "n/a"
        ),
        "Rebalance Frequency": (
            engine.rebalance_freq if hasattr(engine, "rebalance_freq") else "n/a"
        ),
        "Rebalance Count": (
            len(
                [
                    d
                    for d in engine.selection_details
                    if d.get("Rebalance Date") != "SUMMARY"
                ]
            )
            if engine.selection_details
            else 0
        ),
        "Missing Months": len(engine.missing_months),
        "Filtered Tickers": len(engine.filtered_tickers),
        "Run Date": date.today(),
    }
    df = pd.DataFrame(metrics.items(), columns=["Metric", "Value"])
    df.to_excel(writer, sheet_name="Dashboard", index=False)


def create_risk_metrics_sheet(engine, writer):
    """Fügt ein einfaches Risiko-Metrik-Sheet hinzu"""
    if engine.portfolio_value.empty:
        return

    returns = engine.portfolio_value.pct_change().dropna()
    downside_returns = returns[returns < 0]

    sortino = (
        (returns.mean() / downside_returns.std()) * np.sqrt(252)
        if downside_returns.std() > 0
        else 0
    )
    rolling_max = engine.portfolio_value.cummax()
    drawdown = (engine.portfolio_value / rolling_max) - 1
    max_dd = drawdown.min()

    cagr = (
        (engine.portfolio_value.iloc[-1] / engine.start_balance)
        ** (252 / len(engine.portfolio_value))
        - 1
        if len(engine.portfolio_value) > 0
        else 0
    )
    calmar = cagr / abs(max_dd) if max_dd != 0 else None

    var = np.percentile(returns, 5)
    cvar = returns[returns <= var].mean()

    risk_data = {
        "Metric": [
            "Sortino Ratio",
            "Calmar Ratio",
            "Kurtosis",
            "Skewness",
            "VaR (95%)",
            "CVaR (95%)",
        ],
        "Value": [
            round(sortino, 4),
            round(calmar, 4) if calmar else "n/a",
            round(kurtosis(returns), 4),
            round(skew(returns), 4),
            round(var * 100, 4),
            round(cvar * 100, 4),
        ],
    }

    df = pd.DataFrame(risk_data)
    df.to_excel(writer, sheet_name="Risiko", index=False)


def create_drawdown_sheet(engine, writer, top_n=10):
    """Fügt ein Sheet mit den tiefsten Drawdowns hinzu"""
    if engine.portfolio_value.empty:
        return

    df = pd.DataFrame({"Portfolio": engine.portfolio_value})
    df["Peak"] = df["Portfolio"].cummax()
    df["Drawdown"] = df["Portfolio"] / df["Peak"] - 1

    in_drawdown = False
    periods = []
    start = trough = end = None
    peak_val = trough_val = None

    for trade_date, row in df.iterrows():
        if not in_drawdown and row["Drawdown"] < 0:
            in_drawdown = True
            start = date
            peak_val = row["Peak"]
            trough_val = row["Portfolio"]
            trough = date
        elif in_drawdown:
            if row["Portfolio"] < trough_val:
                trough_val = row["Portfolio"]
                trough = date
            if row["Portfolio"] >= peak_val:
                in_drawdown = False
                end = date
                periods.append(
                    {
                        "Start": start.date(),
                        "Trough": trough.date(),
                        "End": end.date(),
                        "Drawdown (%)": round((trough_val / peak_val - 1) * 100, 2),
                        "Duration (Days)": (end - start).days,
                        "Recovery (Days)": (end - trough).days,
                    }
                )

    df_out = pd.DataFrame(periods).sort_values(by="Drawdown (%)").head(top_n)
    df_out.to_excel(writer, sheet_name="Drawdowns", index=False)


def create_trading_costs_sheet(engine, writer):
    """Erstellt eine detaillierte Analyse der Trading-Kosten"""
    if hasattr(engine, "monthly_allocations") and not engine.monthly_allocations.empty:
        # Filter für Einträge mit Trading-Kosten
        try:
            cost_df = engine.monthly_allocations.dropna(subset=["Trading Costs"])

            # Summe der Trading-Kosten pro Rebalance-Datum
            if "Rebalance Date" in cost_df.columns:
                rebalance_costs = (
                    cost_df.groupby("Rebalance Date")["Trading Costs"]
                    .sum()
                    .reset_index()
                )
                rebalance_costs.columns = ["Rebalance Date", "Total Trading Costs"]
                rebalance_costs.to_excel(
                    writer, sheet_name="Trading Costs", index=False
                )

                # Zusätzliche statistische Zusammenfassung
                stats = {
                    "Metric": [
                        "Durchschnittliche Kosten pro Rebalance",
                        "Maximale Kosten bei einem Rebalance",
                        "Minimale Kosten bei einem Rebalance",
                        "Gesamtkosten",
                        "Kosten als % des Startkapitals",
                        "Anzahl der Rebalances",
                    ],
                    "Value": [
                        f"${rebalance_costs['Total Trading Costs'].mean():.2f}",
                        f"${rebalance_costs['Total Trading Costs'].max():.2f}",
                        f"${rebalance_costs['Total Trading Costs'].min():.2f}",
                        f"${engine.total_trading_costs:.2f}",
                        f"{(engine.total_trading_costs / engine.start_balance) * 100:.2f}%",
                        f"{len(rebalance_costs)}",
                    ],
                }

                pd.DataFrame(stats).to_excel(
                    writer, sheet_name="Trading Costs Summary", index=False
                )
        except Exception as e:
            print(f"⚠️ Fehler beim Erstellen des Trading-Kosten-Sheets: {e}")


def create_rebalance_analysis_sheet(engine, writer):
    """Erstellt ein Sheet mit Rebalance-Analysen"""
    if engine.selection_details:
        try:
            # Nur Einträge ohne 'SUMMARY' filtern
            rebalance_details = [
                d
                for d in engine.selection_details
                if d.get("Rebalance Date") != "SUMMARY"
            ]

            # Erstellen eines DataFrames mit relevanten Rebalance-Informationen
            rebalance_df = pd.DataFrame(
                {
                    "Rebalance Date": [
                        d.get("Rebalance Date") for d in rebalance_details
                    ],
                    "Rebalance Frequency": [
                        d.get("Rebalance Frequency", engine.rebalance_freq)
                        for d in rebalance_details
                    ],
                    "Top Universe Size": [
                        d.get("Top Universe Size") for d in rebalance_details
                    ],
                    "Optimization Method": [
                        d.get("Optimization Method") for d in rebalance_details
                    ],
                    "Number of Selected Tickers": [
                        (
                            len(d.get("Selected Tickers", "").split(", "))
                            if d.get("Selected Tickers")
                            else 0
                        )
                        for d in rebalance_details
                    ],
                }
            )

            # Füge einen Spalte für den Zeitabstand hinzu (in Tagen)
            if len(rebalance_df) > 1:
                rebalance_df["Rebalance Date"] = pd.to_datetime(
                    rebalance_df["Rebalance Date"]
                )
                rebalance_df["Days Since Last Rebalance"] = (
                    rebalance_df["Rebalance Date"].diff().dt.days
                )

            rebalance_df.to_excel(writer, sheet_name="Rebalance Analysis", index=False)

            # Summarische Statistiken
            if len(rebalance_df) > 1:
                stats = {
                    "Metric": [
                        "Rebalance Frequency Type",
                        "Actual Average Days Between Rebalances",
                        "Total Number of Rebalances",
                        "First Rebalance Date",
                        "Last Rebalance Date",
                        "Missing Rebalance Months",
                    ],
                    "Value": [
                        engine.rebalance_freq,
                        f"{rebalance_df['Days Since Last Rebalance'].mean():.1f} days",
                        len(rebalance_df),
                        rebalance_df["Rebalance Date"].min().strftime("%Y-%m-%d"),
                        rebalance_df["Rebalance Date"].max().strftime("%Y-%m-%d"),
                        len(engine.missing_months),
                    ],
                }

                pd.DataFrame(stats).to_excel(
                    writer, sheet_name="Rebalance Summary", index=False
                )
        except Exception as e:
            print(f"⚠️ Fehler beim Erstellen des Rebalance-Analyse-Sheets: {e}")
